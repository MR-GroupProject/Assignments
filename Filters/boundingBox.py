from openpyxl import load_workbook
import pymeshlab


wb = load_workbook('../filter.xlsx')
ws = wb['Sheet1']
ws.cell(1, 6).value = 'BBox x'
ws.cell(1, 7).value = 'BBox y'
ws.cell(1, 8).value = 'BBox z'

ws.cell(1, 9).value = 'New BBox x'
ws.cell(1, 10).value = 'New BBox y'
ws.cell(1, 11).value = 'New BBox z'

ms = pymeshlab.MeshSet()

for row in range(2, ws.max_row+1):
    file = ws.cell(row, 1).value
    label = ws.cell(row, 2).value
    ms.load_new_mesh("../LabeledDB_new/" + label + "/" + file)
    ws.cell(row, 6).value = ms.current_mesh().bounding_box().dim_x()
    ws.cell(row, 7).value = ms.current_mesh().bounding_box().dim_y()
    ws.cell(row, 8).value = ms.current_mesh().bounding_box().dim_z()

    ms.compute_matrix_from_scaling_or_normalization(unitflag=True)
    ws.cell(row, 9).value = ms.current_mesh().bounding_box().dim_x()
    ws.cell(row, 10).value = ms.current_mesh().bounding_box().dim_y()
    ws.cell(row, 11).value = ms.current_mesh().bounding_box().dim_z()
    ms.clear()

wb.save('../filter.xlsx')

print('done')
