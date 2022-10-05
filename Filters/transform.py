from openpyxl import load_workbook
import numpy as np
import pymeshlab
import trimesh


def bbox_scaling(ws, ms, start_col=6):
    if ws is not None and ms is not None:
        ws.cell(1, start_col).value = 'BBox x'
        ws.cell(1, start_col + 1).value = 'BBox y'
        ws.cell(1, start_col + 2).value = 'BBox z'

        ws.cell(1, start_col + 3).value = 'New BBox x'
        ws.cell(1, start_col + 4).value = 'New BBox y'
        ws.cell(1, start_col + 5).value = 'New BBox z'

        for row in range(2, ws.max_row + 1):
            file = ws.cell(row, 1).value
            label = ws.cell(row, 2).value
            ms.load_new_mesh("../Remesh/" + label + "/" + file)
            ws.cell(row, start_col).value = ms.current_mesh().bounding_box().dim_x()
            ws.cell(row, start_col + 1).value = ms.current_mesh().bounding_box().dim_y()
            ws.cell(row, start_col + 2).value = ms.current_mesh().bounding_box().dim_z()

            ms.compute_matrix_from_scaling_or_normalization(unitflag=True)
            ws.cell(row, start_col + 3).value = ms.current_mesh().bounding_box().dim_x()
            ws.cell(row, start_col + 4).value = ms.current_mesh().bounding_box().dim_y()
            ws.cell(row, start_col + 5).value = ms.current_mesh().bounding_box().dim_z()
            ms.save_current_mesh("../Remesh/" + label + "/" + file)
            ms.clear()


def translation(ws, ms, start_col=12):
    if ws is not None and ms is not None:
        ws.cell(1, start_col).value = 'Barycenter x'
        ws.cell(1, start_col + 1).value = 'Barycenter y'
        ws.cell(1, start_col + 2).value = 'Barycenter z'
        ws.cell(1, start_col + 3).value = 'Barycenter distance'
        ws.cell(1, start_col + 4).value = 'New barycenter distance'

        for row in range(2, ws.max_row + 1):
            file = ws.cell(row, 1).value
            label = ws.cell(row, 2).value
            ms.load_new_mesh("../Remesh/" + label + "/" + file)
            center = ms.get_geometric_measures().get('barycenter')
            x = center[0]
            y = center[1]
            z = center[2]
            ws.cell(row, start_col).value = x
            ws.cell(row, start_col + 1).value = y
            ws.cell(row, start_col + 2).value = z
            ws.cell(row, start_col + 3).value = x * x + y * y + z * z

            ms.compute_matrix_from_translation(traslmethod=3, neworigin=center)
            new_center = ms.get_geometric_measures().get('barycenter')
            x = new_center[0]
            y = new_center[1]
            z = new_center[2]
            ws.cell(row, start_col + 4).value = x * x + y * y + z * z

            # alignment
            pca(ms)

            ms.save_current_mesh("../Remesh/" + label + "/" + file)
            # flip test
            flip_mesh = trimesh.load_mesh("../Remesh/" + label + "/" + file)
            tri_center = flip_mesh.triangles_center
            f0, f1, f2 = 0, 0, 0

            for point in tri_center:
                f0 += point[0] * point[0] * np.sign(point[0])
                f1 += point[1] * point[1] * np.sign(point[1])
                f2 += point[2] * point[2] * np.sign(point[2])

            if f0 >= 0 and f1 >= 0 and f2 >= 0:
                ms.clear()
                continue
            else:
                print(file)
                if f0 < 0:
                    ms.apply_matrix_flip_or_swap_axis(flipx=True)
                if f1 < 0:
                    ms.apply_matrix_flip_or_swap_axis(flipy=True)
                if f2 < 0:
                    ms.apply_matrix_flip_or_swap_axis(flipz=True)
                if f0 * f1 * f2 < 0:
                    ms.meshing_invert_face_orientation()
                ms.save_current_mesh("../Remesh/" + label + "/" + file)
                ms.clear()


def pca(ms):
    mesh_np = ms.current_mesh().vertex_matrix()  # get mesh matrix
    # face_matrix = ms.face_matrix()
    cov_np = np.cov(mesh_np, rowvar=False)  # calculate the covariance matrix
    feature_val, feature_vect = np.linalg.eig(np.mat(cov_np))

    feature_val_index = np.argsort(feature_val)
    number_feature_val_index = feature_val_index[-1:-4:-1]
    number_feature_vect = feature_vect[:, number_feature_val_index]
    number_feature_vect = number_feature_vect.T

    e1 = number_feature_vect[0]
    e2 = number_feature_vect[1]
    e3 = np.cross(e1, e2)
    m = np.array([e1, e2, e3])
    m = np.mat(m)
    # result = np.dot(mesh_np, m.T)
    m = np.insert(m, 3, [0, 0, 0], 0)
    m = np.insert(m, 3, [0, 0, 0, 1], 1)

    ms.set_matrix(transformmatrix=m)  # transform the mesh


wb = load_workbook('../filter.xlsx')
sheet = wb['Sheet1']
mesh_set = pymeshlab.MeshSet()

translation(sheet, mesh_set)
#bbox_scaling(sheet, mesh_set)
wb.save('../filter.xlsx')
