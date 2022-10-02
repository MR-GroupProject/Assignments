from openpyxl import load_workbook
import pymeshlab


wb = load_workbook('../filter.xlsx')
ws = wb['Sheet1']
ws.cell(1, 12).value = 'Barycenter x'
ws.cell(1, 13).value = 'Barycenter y'
ws.cell(1, 14).value = 'Barycenter z'
ws.cell(1, 15).value = 'Barycenter distance'
ws.cell(1, 16).value = 'New barycenter distance'


ms = pymeshlab.MeshSet()

for row in range(2, ws.max_row+1):
    file = ws.cell(row, 1).value
    label = ws.cell(row, 2).value
    ms.load_new_mesh("../LabeledDB_new/" + label + "/" + file)
    center = ms.get_geometric_measures().get('barycenter')
    x = center[0]
    y = center[1]
    z = center[2]
    ws.cell(row, 12).value = x
    ws.cell(row, 13).value = y
    ws.cell(row, 14).value = z
    ws.cell(row, 15).value = x*x + y*y + z*z

    ms.compute_matrix_from_translation(traslmethod=3, neworigin=center)
    new_center = ms.get_geometric_measures().get('barycenter')
    x = new_center[0]
    y = new_center[1]
    z = new_center[2]
    ws.cell(row, 16).value = x*x + y*y + z*z

    ms.clear()

wb.save('../filter.xlsx')

print('done')