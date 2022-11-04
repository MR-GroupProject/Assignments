from openpyxl import load_workbook
import pymeshlab
import os
from tools import normalize


def resampling(ws, ms, start_col=17):
    if ws is not None and ms is not None:
        ws.cell(1, start_col).value = 'New v'
        ws.cell(1, start_col + 1).value = 'New f'
        for row in range(2, ws.max_row + 1):
            file = ws.cell(row, 1).value
            label = ws.cell(row, 2).value
            ms.load_new_mesh("../LabeledDB_new/" + label + "/" + file)

            if ws.cell(row, 4).value > 15000 or ws.cell(row, 4).value < 10000:
                normalize.re_mesh(ms=ms)
                ws.cell(row, start_col).value = ms.current_mesh().vertex_number()
                ws.cell(row, start_col + 1).value = ms.current_mesh().face_number()
                print("done:" + file)
            else:
                ws.cell(row, start_col).value = ws.cell(row, 3).value
                ws.cell(row, start_col + 1).value = ws.cell(row, 4).value

            normalize.clean_ms(ms)
            ms.save_current_mesh("../Remesh/" + label + "/" + file)
            ms.clear()


wb = load_workbook('../filter.xlsx')
sheet = wb['Sheet1']
mesh = pymeshlab.MeshSet()

# create dictionary to store mesh files after re meshing
if not os.path.exists('../Remesh'):
    os.mkdir('../Remesh/')
for r in range(2, sheet.max_row + 1):
    label = sheet.cell(r, 2).value
    if not os.path.exists('../Remesh/' + label):
        os.mkdir('../Remesh/' + label)

resampling(sheet, mesh)
wb.save('../filter.xlsx')
