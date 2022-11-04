from openpyxl import load_workbook
import numpy as np
import pymeshlab
import trimesh
from tools import normalize


def bbox_scaling(ws, ms, start_col=6):
    if ws is not None and ms is not None:
        ws.cell(1, start_col).value = 'BBox x'
        ws.cell(1, start_col + 1).value = 'BBox y'
        ws.cell(1, start_col + 2).value = 'BBox z'

        ws.cell(1, start_col + 3).value = 'New BBox x'
        ws.cell(1, start_col + 4).value = 'New BBox y'
        ws.cell(1, start_col + 5).value = 'New BBox z'

        for row in range(2, ws.max_row + 1):
            file = ws.cell(row, 1).value
            label = ws.cell(row, 2).value
            ms.load_new_mesh("../Remesh/" + label + "/" + file)
            ws.cell(row, start_col).value = ms.current_mesh().bounding_box().dim_x()
            ws.cell(row, start_col + 1).value = ms.current_mesh().bounding_box().dim_y()
            ws.cell(row, start_col + 2).value = ms.current_mesh().bounding_box().dim_z()

            ms.compute_matrix_from_scaling_or_normalization(unitflag=True)
            ws.cell(row, start_col + 3).value = ms.current_mesh().bounding_box().dim_x()
            ws.cell(row, start_col + 4).value = ms.current_mesh().bounding_box().dim_y()
            ws.cell(row, start_col + 5).value = ms.current_mesh().bounding_box().dim_z()
            ms.save_current_mesh("../Remesh/" + label + "/" + file)
            # print("done: " + label)
            ms.clear()


def translation(ws, ms, start_col=12):
    if ws is not None and ms is not None:
        ws.cell(1, start_col).value = 'Barycenter x'
        ws.cell(1, start_col + 1).value = 'Barycenter y'
        ws.cell(1, start_col + 2).value = 'Barycenter z'
        ws.cell(1, start_col + 3).value = 'Barycenter distance'
        ws.cell(1, start_col + 4).value = 'New barycenter distance'

        for row in range(2, ws.max_row + 1):
            file = ws.cell(row, 1).value
            label = ws.cell(row, 2).value
            ms.load_new_mesh("../Remesh/" + label + "/" + file)
            center = ms.get_geometric_measures().get('barycenter')
            x = center[0]
            y = center[1]
            z = center[2]
            ws.cell(row, start_col).value = x
            ws.cell(row, start_col + 1).value = y
            ws.cell(row, start_col + 2).value = z
            ws.cell(row, start_col + 3).value = x * x + y * y + z * z

            ms.compute_matrix_from_translation(traslmethod=3, neworigin=center)
            new_center = ms.get_geometric_measures().get('barycenter')
            x = new_center[0]
            y = new_center[1]
            z = new_center[2]
            ws.cell(row, start_col + 4).value = x * x + y * y + z * z

            # alignment
            normalize.pca(ms)

            ms.save_current_mesh("../Remesh/" + label + "/" + file)
            # flip test
            flip_label = normalize.flip(ms, "../Remesh/" + label + "/" + file)
            if flip_label == 0:
                ms.clear()
                continue

            else:
                ms.save_current_mesh("../Remesh/" + label + "/" + file)
                ms.clear()


wb = load_workbook('../filter.xlsx')
sheet = wb['Sheet1']
mesh_set = pymeshlab.MeshSet()

translation(sheet, mesh_set)
bbox_scaling(sheet, mesh_set)
wb.save('../filter.xlsx')
