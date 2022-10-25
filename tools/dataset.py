import features as ft
import xlsxwriter
from openpyxl import load_workbook
import os


# create data sheet
def create_table(path, labels=None, sheet=None):
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet()
    if sheet is not None:
        worksheet.set_vba_name(sheet)
    if labels is not None:
        worksheet.write_row('A1', labels)
    return workbook, worksheet


def write_data(data, worksheet, start_row=0, start_column=0):
    for i in range(len(data)):
        worksheet.write_row(start_row + i, start_column, data[i])


def get_all_data(filepath, sheet='Sheet1'):
    wb = load_workbook(filepath)
    ws = wb[sheet]
    result = []

    for row in ws.rows:
        row_cells = []
        for cell in row:
            row_cells.append(cell.value)
        result.append(row_cells)

    return result
